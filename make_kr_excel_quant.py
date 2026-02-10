import random, time
import os
import requests
import pandas as pd
import math
import concurrent.futures
from pathlib import Path
from dotenv import load_dotenv
from tqdm import tqdm
from datetime import datetime
from modules.naver_upjong_quant import fetch_stock_info_quant_API
from openpyxl import load_workbook
from openpyxl.styles import Font

# API 기본 URL
kospi_url = "https://m.stock.naver.com/api/stocks/marketValue/KOSPI"
kosdaq_url = "https://m.stock.naver.com/api/stocks/marketValue/KOSDAQ"
# 최대 스레드 수
max_workers = 4

# 요청 헤더
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}

def ensure_directory(path):
    """폴더가 존재하지 않으면 생성"""
    if not os.path.exists(path):
        os.makedirs(path)

def fetch_all_stocks(base_url, market_name):
    params = {"page": 1, "pageSize": 100}
    response = requests.get(base_url, headers=headers, params=params)
    
    if response.status_code != 200:
        print(f"{market_name} 첫 페이지 요청 실패: {response.status_code}")
        return []
    
    data = response.json()
    total_count = data.get("totalCount", 0)
    if total_count == 0:
        print(f"{market_name}에서 totalCount를 가져올 수 없습니다.")
        return []
    
    page_size = 100
    total_pages = math.ceil(total_count / page_size)
    print(f"{market_name} - 총 종목 수: {total_count}, 총 페이지 수: {total_pages}")
    
    all_stocks = []
    etf_etn_stocks = []
    for page in tqdm(range(1, total_pages + 1), desc=f"Fetching {market_name}"):
        params = {"page": page, "pageSize": page_size}
        response = requests.get(base_url, headers=headers, params=params)
        
        if response.status_code == 200:
            page_data = response.json()
            stocks = page_data.get("stocks", [])
            for stock in stocks:
                stock_name = stock.get("stockName", "")
                if ('호스팩' in stock_name) or (stock_name.endswith('스팩') and '호' in stock_name):
                    print(f"스팩 종목으로 추정되는 '{stock_name}'을(를) 건너뜁니다.")
                    continue

                stock_info = {
                    "종목코드": stock.get("itemCode", ""),
                    "종목명": stock.get("stockName", ""),
                    "시가총액(억)": int(stock.get("marketValue", 0).replace(",", "")),
                }
                if stock.get("stockEndType") in ["etf", "etn"]:
                    etf_etn_stocks.append(stock_info)
                else:
                    all_stocks.append(stock_info)
        else:
            print(f"{market_name} 페이지 {page} 요청 실패: {response.status_code}")
    
    return all_stocks, etf_etn_stocks

def fetch_quant_data(stock_info):
    stock_code = stock_info["종목코드"]
    url = f"https://m.stock.naver.com/item/main.nhn#/stocks/{stock_code}/total"
    quant_data = fetch_stock_info_quant_API(stock_code, url)
    time.sleep(random.uniform(1.0, 2.0))
    return quant_data if quant_data else {}

def add_quant_data(df, market_name):
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        quant_data_list = list(tqdm(executor.map(fetch_quant_data, df.to_dict('records')), total=len(df), desc=f"Fetching Quant Data for {market_name}"))
    
    quant_df = pd.DataFrame(quant_data_list)
    df = pd.concat([df, quant_df], axis=1)
    df = df.sort_values(by="시가총액(억)", ascending=False)
    df = df.loc[:, ~df.columns.duplicated()]
    return df

def send_to_telegram():
    # .env 파일 로드
    env_path = Path('/home/ubuntu/dev/telegram-stock-info-noti-bot/.env')
    load_dotenv(dotenv_path=env_path)

    # 폴더 경로 설정
    FOLDER_PATH = "/home/ubuntu/dev/telegram-stock-info-bot"
    # FOLDER_PATH = "/Users/seunghoonshin/dev/telegram-stock-info-bot"
    SEND_FOLDER = os.path.join(FOLDER_PATH, "send")

    # 최신 엑셀 파일 찾기 (KR_stock_screening_YYMMDD.xlsx 형식)
    files = sorted(
        [f for f in os.listdir(FOLDER_PATH) if f.startswith("KR_stock_screening_") and f.endswith(".xlsx")],
        reverse=True
    )
    
    if not files:
        print(f"오류: 전송할 파일을 찾을 수 없습니다. ({FOLDER_PATH}/KR_stock_screening_*.xlsx)")
        return False

    FILE_PATH = os.path.join(FOLDER_PATH, files[0])
    FILE_NAME = os.path.basename(FILE_PATH)

    # 메시지 설정
    MESSAGE = f"📊 주식 스크리닝 결과 파일 전송: {FILE_NAME}"

    # Telegram API 요청
    url = f"https://api.telegram.org/bot{os.getenv('TELEGRAM_BOT_TOKEN_PROD')}/sendDocument"
    data = {
        'chat_id': os.getenv('TELEGRAM_CHANNEL_ID_REPORT_ALARM'),
        'caption': MESSAGE
    }
    
    with open(FILE_PATH, 'rb') as file:
        files = {'document': file}
        response = requests.post(url, data=data, files=files)
    
    # 응답 출력
    print(f"Response: {response.text}")

    # send 폴더가 없으면 생성
    os.makedirs(SEND_FOLDER, exist_ok=True)

    # 전송한 파일 이동
    new_file_path = os.path.join(SEND_FOLDER, FILE_NAME)
    os.rename(FILE_PATH, new_file_path)
    
    print(f"파일이 전송 후 이동되었습니다: {new_file_path}")
    return True

def main():
    today_date = datetime.today().strftime('%y%m%d')
    send_folder = "send"
    ensure_directory(send_folder)
    
    file_name = f"KR_stock_screening_{today_date}.xlsx"
    file_path = os.path.join(send_folder, file_name)
    
    # 1. send_folder에 해당일자 파일이 있는 경우 정상 종료
    if os.path.exists(file_path):
        print(f"{file_name} 파일이 이미 존재합니다. 프로그램을 종료합니다.")
        return
    
    # 2. send_folder에 파일이 없고, send_to_telegram()로 보낼 파일이 없는 경우
    if not send_to_telegram():  # 최초 전송 시도 후 파일이 없으면 데이터 생성
        kospi_stocks, kospi_etf_etn = fetch_all_stocks(kospi_url, "KOSPI")
        kosdaq_stocks, kosdaq_etf_etn = fetch_all_stocks(kosdaq_url, "KOSDAQ")
        
        df_kospi = pd.DataFrame(kospi_stocks)
        df_kosdaq = pd.DataFrame(kosdaq_stocks)
        df_etf_etn = pd.DataFrame(kospi_etf_etn + kosdaq_etf_etn)
        
        df_kospi = add_quant_data(df_kospi, "KOSPI")
        df_kosdaq = add_quant_data(df_kosdaq, "KOSDAQ")
        df_etf_etn = add_quant_data(df_etf_etn, "ETF_ETN")
        
        # send_folder가 아닌 기본 경로에 저장
        base_path = "/home/ubuntu/dev/telegram-stock-info-bot"
        file_path = os.path.join(base_path, file_name)
        
        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            df_kospi.to_excel(writer, sheet_name="KOSPI", index=False)
            df_kosdaq.to_excel(writer, sheet_name="KOSDAQ", index=False)
            df_etf_etn.to_excel(writer, sheet_name="ETF_ETN", index=False)
        
        print(f"\n데이터가 '{file_name}' 파일에 저장되었습니다. (시트: KOSPI, KOSDAQ, ETF_ETN)")
        
        # 엑셀 서식 적용
        beautify_excel(file_path)  # 엑셀 서식 적용
        # 생성된 파일을 Telegram으로 전송
        send_to_telegram()
    
    else:
        print("이미 처리된 파일이 Telegram으로 전송되었습니다.")


def beautify_excel(file_path):
    """
    생성된 엑셀 파일을 불러와서 서식을 적용하는 함수
    1. 헤더에 필터 적용
    2. B열(종목명) 너비 조정 및 폰트 강조
    """
    # 엑셀 파일 불러오기
    wb = load_workbook(file_path)
    
    # 모든 시트에 대해 반복
    for ws in wb.worksheets:
        # 1. 최상단 필터 설정 (데이터가 있는 전체 영역에 적용)
        ws.auto_filter.ref = ws.dimensions

        # 2. B열(종목명) 꾸미기
        target_column = 'B'  # B열
        
        # 2-1. 너비 늘리기 (기본값 약 8 -> 25로 변경)
        ws.column_dimensions[target_column].width = 25

        # 2-2. 폰트 설정 (Bold, 사이즈 12)
        # 맑은 고딕(Mac이면 AppleGothic), 크기 16, 굵게
        b_col_font = Font(bold=True, size=16, name='Malgun Gothic')

        # 헤더(1행)는 건너뛰고 2행부터 적용
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.font = b_col_font

    # 저장
    wb.save(file_path)
    print(f"✅ 서식 적용 완료: {file_path}")
    
if __name__ == '__main__':
    main()