import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime
import argparse

# 현재 스크립트의 상위 디렉터리를 모듈 경로에 추가
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from modules.naver_upjong_quant import fetch_upjong_list_API, fetch_stock_info_in_upjong, fetch_stock_info_quant_API

# Define the folder path
EXCEL_FOLDER_PATH = 'excel/'  # Adjust this to your actual folder path if needed

# Argument parser 설정
parser = argparse.ArgumentParser(description='엑셀 파일 생성 옵션')
parser.add_argument('--sheet_type', choices=['업종별', '전체'], default='업종별', help='생성할 시트 유형')
args = parser.parse_args()

# 업종별 시트와 전체 시트 선택 변수
SHEET_TYPE = args.sheet_type

# 업종 검색 처리
upjong_list = fetch_upjong_list_API('KOR')
upjong_map = {업종명: (등락률, 링크) for 업종명, 등락률, 링크 in upjong_list}

# Ensure the folder exists
if not os.path.exists(EXCEL_FOLDER_PATH):
    os.makedirs(EXCEL_FOLDER_PATH)

today_date = datetime.today().strftime('%y%m%d')
# 엑셀 파일명 생성
excel_file_name = os.path.join(EXCEL_FOLDER_PATH, f'네이버전체업종_naver_quant_{today_date}.xlsx')

# 새 엑셀 파일 생성
wb = Workbook()
if SHEET_TYPE == '전체':
    wb.remove(wb.active)  # '전체' 시트에 대해 기본 시트를 제거합니다.

def process_stock_info(stock_info):
    all_quant_data = []
    for 종목명, _, _, _, 종목링크 in stock_info:
        stock_code = 종목링크.split('=')[-1]
        quant_data = fetch_stock_info_quant_API(stock_code)
        if quant_data:
            all_quant_data.append(quant_data)
    return all_quant_data

def save_to_excel(df, sheet_name):
    # 시트 이름에 따라 새 시트 생성
    if sheet_name == '전체':
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # 데이터프레임을 시트에 작성
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # 필터 기능 추가
    start_row, start_col = 1, 1
    end_row = ws.max_row
    end_col = ws.max_column
    cell_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
    ws.auto_filter.ref = cell_range

    # '네이버url' 열의 하이퍼링크 처리
    if '네이버url' in df.columns:
        nav_url_col_index = df.columns.get_loc('네이버url') + 1  # 1-based index
        for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=nav_url_col_index, max_col=nav_url_col_index):
            for cell in row:
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")

if SHEET_TYPE == '업종별':
    # 업종별 시트 생성
    for 업종명, (등락률, 링크) in upjong_map.items():
        print(f"=================업종명: {업종명}, 등락률: {등락률}=================")
        if 업종명 == '기타':  # ETN & ETF는 건너뛰기
            continue

        stock_info = fetch_stock_info_in_upjong(링크)
        if stock_info:
            all_quant_data = process_stock_info(stock_info)
            if all_quant_data:
                df = pd.DataFrame(all_quant_data)
                save_to_excel(df, 업종명)

elif SHEET_TYPE == '전체':
    # 전체 시트 생성
    all_quant_data = []
    for 업종명, (등락률, 링크) in upjong_map.items():
        print(f"=================업종명: {업종명}, 등락률: {등락률}=================")
        if 업종명 == '기타':  # ETN & ETF는 건너뛰기
            continue

        stock_info = fetch_stock_info_in_upjong(링크)
        if stock_info:
            all_quant_data.extend(process_stock_info(stock_info))

    if all_quant_data:
        df = pd.DataFrame(all_quant_data)
        save_to_excel(df, '전체')

# 엑셀 파일 저장
wb.save(excel_file_name)
wb.close()

print(f'퀀트 정보가 {excel_file_name} 파일에 저장되었습니다.')
