from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# 엑셀 셀 주소 변환 함수
def number_to_coordinate(rc):
    row_idx, col_idx = rc[0], rc[1]
    col_string = get_column_letter(col_idx)
    return f'{col_string}{row_idx}'

def adjust_column_width(ws):
    """첫 번째 행을 기준으로 각 열의 너비를 자동으로 조정합니다."""
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            adjusted_width = len(str(cell_value)) * 4
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width

def apply_auto_filter(ws):
    """데이터가 있는 범위를 자동 필터로 설정합니다."""
    start_row, start_col = 1, 1
    end_row = ws.max_row
    end_col = ws.max_column
    cell_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
    ws.auto_filter.ref = cell_range

def apply_hyperlinks(ws, df):
    """'네이버url' 열에 하이퍼링크를 적용합니다."""
    url_column_name = '네이버url'
    if url_column_name in df.columns:
        url_col_index = df.columns.get_loc(url_column_name) + 1  # 1-based index
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=url_col_index, max_col=url_col_index):
            for cell in row:
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")

def process_excel_file(file_name, df):
    """
    엑셀 파일의 모든 시트에서 열 너비 조정, 자동 필터 적용, 하이퍼링크 처리를 수행합니다.

    Parameters:
    file_name (str): 처리할 엑셀 파일의 이름
    df (pandas.DataFrame): 하이퍼링크를 적용할 데이터가 포함된 DataFrame

    Returns:
    None
    """
    # 엑셀 파일을 로드합니다.
    wb = load_workbook(file_name)
    
    # 모든 시트를 순회하며 처리합니다.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 열 너비를 조정합니다.
        adjust_column_width(ws)
        
        # 자동 필터를 적용합니다.
        apply_auto_filter(ws)
        
        # 하이퍼링크를 적용합니다.
        apply_hyperlinks(ws, df)
    
    # 변경된 내용을 저장합니다.
    wb.save(file_name)
    
    # 워크북을 닫습니다.
    wb.close()

