import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink

# 업종 페이지 URL
base_upjong_url = 'https://finance.naver.com/sise/sise_group.naver?type=upjong'

# HTTP 헤더 설정
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

def fetch_upjong_list():
    response = requests.get(base_upjong_url, headers=headers)
    response.encoding = 'euc-kr'
    soup = BeautifulSoup(response.text, 'html.parser')

    table = soup.find('table', {'class': 'type_1'})
    if not table:
        raise ValueError("업종 목록 테이블을 찾을 수 없습니다.")
    
    rows = table.find_all('tr')[2:]
    data = []
    for row in rows:
        cols = row.find_all('td')
        if len(cols) >= 2:
            업종명 = cols[0].get_text(strip=True)
            등락률 = cols[1].get_text(strip=True)
            링크 = cols[0].find('a')['href']
            data.append((업종명, 등락률, 링크))
    
    return data

def fetch_stock_info(upjong_link):
    base_url = 'https://finance.naver.com'
    full_url = base_url + upjong_link

    response = requests.get(full_url, headers=headers)
    response.encoding = 'euc-kr'
    soup = BeautifulSoup(response.text, 'html.parser')

    table = soup.find('table', {'class': 'type_5'})
    if not table:
        raise ValueError(f'종목 정보를 찾을 수 없습니다: {full_url}')
    
    rows = table.find_all('tr')[1:]
    stock_data = []
    for row in rows:
        cols = row.find_all('td')
        if len(cols) >= 10:
            종목명 = cols[0].get_text(strip=True)
            현재가 = cols[1].get_text(strip=True)
            전일비_raw = cols[2].get_text(strip=True)
            등락률 = cols[3].get_text(strip=True)
            
            if '상승' in 전일비_raw:
                전일비 = '+' + 전일비_raw.replace('상승', '').strip()
            elif '하락' in 전일비_raw:
                전일비 = '-' + 전일비_raw.replace('하락', '').strip()
            else:
                전일비 = 전일비_raw

            link_tag = cols[0].find('a')
            link = base_url + link_tag['href'] if link_tag and 'href' in link_tag.attrs else 'N/A'
            
            stock_data.append((종목명, 현재가, 전일비, 등락률, link))
    
    return stock_data

def fetch_stock_info_quant(stock_code):
    url = f'https://finance.naver.com/item/main.naver?code={stock_code}'
    response = requests.get(url, headers=headers)
    response.encoding = 'euc-kr'
    soup = BeautifulSoup(response.text, 'html.parser')

    stock_name_tag = soup.find('div', {'class': 'wrap_company'}).find('h2')
    stock_name = stock_name_tag.get_text(strip=True).split(' ')[0] if stock_name_tag else 'N/A'

    info_section = soup.select_one('#tab_con1 > div:nth-child(5)')
    data = {'종목명': stock_name}

    try:
        current_price_tag = soup.select_one('#middle > dl > dd:nth-child(5)')
        current_price_text = current_price_tag.get_text(strip=True) if current_price_tag else 'N/A'

        pattern = re.compile(r'현재가\s+([\d,]+)\s+전일대비\s+(상승|하락)\s+([\d,]+)\s*(\d+\.\d+)\s*퍼센트')
        match = pattern.search(current_price_text)
        if match:
            data['현재가'] = int(match.group(1).replace(',', ''))
            data['전일비'] = f"{'-' if match.group(2) == '하락' else '+'}{match.group(3).replace(',', '')}"
            data['등락률'] = match.group(4)
        else:
            data['현재가'], data['전일비'], data['등락률'] = 'N/A', 'N/A', 'N/A'

        per_value = info_section.select_one('tr:nth-of-type(1) > td')
        data['PER'] = per_value.get_text(strip=True).split('l')[0].split('배')[0] if per_value else 'N/A'

        est_eps_value = soup.select_one('#_cns_eps')
        est_eps_text = est_eps_value.get_text(strip=True).split('|')[0] if est_eps_value else 'N/A'
        if est_eps_text != 'N/A':
            try:
                est_eps_value = float(est_eps_text.replace(',', ''))
                data['fwdPER'] = round(data['현재가'] / est_eps_value, 2)
            except ValueError:
                data['fwdPER'] = 'N/A'
        else:
            data['fwdPER'] = 'N/A'

        pbr_value = info_section.select_one('tr:nth-of-type(3) > td')
        data['PBR'] = pbr_value.get_text(strip=True).split('배')[0] if pbr_value else 'N/A'

        dividend_yield_value = info_section.select_one('tr:nth-of-type(4) > td')
        data['배당수익률'] = dividend_yield_value.get_text(strip=True).split('%')[0] if dividend_yield_value else 'N/A'
    except Exception as e:
        print(f"데이터 처리 중 오류 발생: {e}")
        data = {}

    data['네이버url'] = url
    return data

def save_to_excel(stock_data, filename='stock_data.xlsx'):
    df = pd.DataFrame(stock_data)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='종목 정보')
        writer.save()

    # Load the workbook to add hyperlinks
    wb = load_workbook(filename)
    ws = wb['종목 정보']

    # Add hyperlinks to the '네이버url' column
    for row in range(2, len(stock_data) + 2):
        url = ws.cell(row=row, column=5).value  # Assuming '네이버url' is in the 5th column
        if url and url != 'N/A':
            ws.cell(row=row, column=5).hyperlink = Hyperlink(ref="", target=url)

    wb.save(filename)

def main():
    upjong_list = fetch_upjong_list()

    all_stock_data = []
    for 업종명, 등락률, 링크 in upjong_list:
        stock_info = fetch_stock_info(링크)
        all_stock_data.extend(stock_info)
    
    save_to_excel(all_stock_data)

if __name__ == "__main__":
    main()
