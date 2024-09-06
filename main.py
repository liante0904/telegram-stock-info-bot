from telegram import Update, BotCommand, InputFile
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, CallbackQueryHandler, CallbackContext
import os
import pandas as pd
import asyncio
import re
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from dotenv import load_dotenv
from module.naver_upjong_quant import fetch_upjong_list_API, fetch_stock_info_in_upjong, fetch_stock_info_quant_API
from module.naver_stock_util import search_stock
from module.chart import CHART_DIR
from module.recent_searches import load_recent_searches, show_recent_searches
from handler.report_handler import process_report_request, previous_search, process_selected_stock_for_report
from handler.chart_handler import process_selected_stock_for_chart, process_generate_chart_stock_list
from handler.quant_handler import process_selected_stock_for_quant
from handler.upjong_handler import show_upjong_list
from datetime import datetime, timedelta


# JSON 파일 경로
KEYWORD_FILE_PATH = 'report_alert_keyword.json'
# Define the folder path
CSV_FOLDER_PATH = 'csv/'  # Adjust this to your actual folder path if needed
EXCEL_FOLDER_PATH = 'excel/'  # Adjust this to your actual folder path if needed

async def generate_chart(update: Update, context: CallbackContext) -> None:
    chat_id = update.effective_chat.id
    await context.bot.send_message(chat_id=chat_id, text='수급오실레이터 차트 생성입니다. \n\n 종목명 혹은 종목코드를 입력하세요. \n 쉼표(,) 혹은 여러줄로 입력하면 다중생성이 가능합니다. \n 종목코드로 입력시 더 빠름')
    context.user_data['next_command'] = 'generate_chart'

async def stock_quant(update: Update, context: CallbackContext) -> None:
    chat_id = update.effective_chat.id
    await context.bot.send_message(chat_id=chat_id, text='종목 & ETF 퀀트입니다. \n\n 종목명 혹은 종목코드를 입력하세요.(ETF가능) \n 쉼표(,) 혹은 여러줄로 입력하면 다중생성이 가능합니다. \n 종목코드로 입력시 더 빠름')
    context.user_data['next_command'] = 'stock_quant'

async def excel_quant(update: Update, context: CallbackContext) -> None:
    chat_id = update.effective_chat.id
    await context.bot.send_message(chat_id=chat_id, text='엑셀 퀀트입니다. \n\n업종, 종목 퀀트 엑셀 파일을 보내면 \n*최근 거래일 데이터로 갱신*합니다. \n"네이버url, 종목코드, 종목명" 중 하나를 기준으로 합니다. \n*파일 전송시 caption값을 넣으면 파일명*을 바꿔 보내줍니다.', parse_mode='Markdown')
    context.user_data['next_command'] = 'excel_quant'

async def search_report(update: Update, context: CallbackContext) -> None:
    chat_id = update.effective_chat.id
    await context.bot.send_message(chat_id=chat_id, text='레포트를 검색할 종목명을 입력하세요.')
    context.user_data['next_command'] = 'search_report'

async def report_alert_keyword(update: Update, context: CallbackContext) -> None:
    chat_id = update.effective_chat.id
    user_id = str(update.effective_user.id)
    
    # 현재 저장된 키워드 로드
    all_keywords = load_alert_keywords()
    current_keywords = [keyword['keyword'] for keyword in all_keywords.get(user_id, [])]

    # 사용자에게 현재 저장된 키워드를 보여주고 입력 요청
    if current_keywords:
        keyword_text = '\n'.join([f"- {keyword}" for keyword in current_keywords])
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"현재 저장된 알림 키워드:\n{keyword_text}\n\n새로운 키워드를 쉼표(,) 또는 하이픈(-)으로 구분하여 입력해주세요. \n\n '키워드 삭제' 를 하면 전체 키워드가 삭제 됩니다."
        )
    else:
        await context.bot.send_message(
            chat_id=chat_id,
            text='현재 저장된 알림 키워드가 없습니다. 새로운 키워드를 쉼표(,) 또는 하이픈(-)으로 구분하여 입력해주세요.'
        )

    # 다음 명령어 상태 설정
    context.user_data['next_command'] = 'report_alert_keyword'

# JSON 파일에서 사용자 알림 키워드를 불러오는 함수
def load_alert_keywords():
    if os.path.exists(KEYWORD_FILE_PATH):
        with open(KEYWORD_FILE_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    return {}

# JSON 파일에 사용자 알림 키워드를 저장하는 함수
def save_alert_keywords(keywords):
    with open(KEYWORD_FILE_PATH, 'w', encoding='utf-8') as file:
        json.dump(keywords, file, ensure_ascii=False, indent=4)

async def route_command_based_on_user_input(update: Update, context: CallbackContext) -> None:
    """
    Handles the user's input after a command has been selected. This function 
    processes the input (stock name or code) and routes the request to the 
    appropriate processing function based on the selected command.

    Args:
        update (Update): The update object containing the user input and callback query.
        context (CallbackContext): The context object containing user data and other information.

    Processes the following commands based on user input:
        - 'generate_chart': Calls `process_selected_stock_for_chart` to generate a chart for the selected stock.
        - 'search_report': Calls `process_selected_stock_for_report` to generate a report for the selected stock.
        - 'stock_quant': Calls `process_selected_stock_for_quant` to provide stock quant analysis for the selected stock.
    """
    query = update.callback_query
    await query.answer()
    selected_code = query.data
    results = context.user_data.get('search_results', [])
    next_command = context.user_data.get('next_command')

    for result in results:
        if result['code'] == selected_code:
            stock_name, stock_code = result['name'], result['code']
            if next_command == 'generate_chart':
                await process_selected_stock_for_chart(update, context, stock_name, stock_code)
            elif next_command == 'search_report':
                await process_selected_stock_for_report(update, context, stock_name, stock_code)
            elif next_command == 'stock_quant':
                await process_selected_stock_for_quant(update, context, stock_name, stock_code)

async def set_commands(bot):
    commands = [
        BotCommand("generate_chart", "수급오실레이터 차트"),
        BotCommand("recent", "최근 검색 종목"),
        BotCommand("search_report", "레포트 검색기"),
        BotCommand("naver_upjong_quant", "네이버 업종퀀트"),  # 새로 추가된 명령어
        BotCommand("stock_quant", "종목 퀀트"),  # 새로 추가된 명령어
        BotCommand("excel_quant", "엑셀 퀀트"),  # 새로 추가된 명령어
        BotCommand("report_alert_keyword", "레포트 알림 키워드 설정")
    ]
    await bot.set_my_commands(commands)

async def handle_message(update: Update, context: CallbackContext) -> None:
    user_id = str(update.effective_user.id)
    user_input = update.message.text
    chat_id = update.effective_chat.id
    next_command = context.user_data.get('next_command')
    print('next_command', next_command)
    try:
        if next_command == 'report_alert_keyword':
            if user_input.lower() == '키워드 삭제':
                all_keywords = load_alert_keywords()
                
                # 사용자 아이디에 해당하는 키워드 리스트를 빈 리스트로 설정
                if user_id in all_keywords:
                    all_keywords[user_id] = []  # 해당 사용자의 모든 키워드를 삭제
                    save_alert_keywords(all_keywords)
                    
                    await context.bot.send_message(
                        chat_id=chat_id,
                        text='모든 알림 키워드가 삭제되었습니다.'
                    )
                else:
                    await context.bot.send_message(
                        chat_id=chat_id,
                        text='삭제할 키워드가 없습니다.'
                    )
            else:
                # 키워드 알림 처리
                keywords = [keyword.strip() for keyword in re.split('[,-]', user_input) if keyword.strip()]
                unique_keywords = set(keywords)
                all_keywords = load_alert_keywords()

                if user_id not in all_keywords:
                    all_keywords[user_id] = []

                existing_keywords = {entry['keyword'] for entry in all_keywords.get(user_id, [])}
                new_keywords = [{'keyword': keyword, 'code': '', 'timestamp': datetime.now().isoformat()} for keyword in unique_keywords if keyword not in existing_keywords]
                all_keywords[user_id].extend(new_keywords)
                unique_user_keywords = {entry['keyword']: entry for entry in all_keywords[user_id]}
                all_keywords[user_id] = list(unique_user_keywords.values())

                save_alert_keywords(all_keywords)

                context.user_data['next_command'] = None
                updated_keywords = [keyword['keyword'] for keyword in all_keywords[user_id]]
                updated_keywords_text = '\n'.join([f"- {keyword}" for keyword in updated_keywords])
                await context.bot.send_message(
                    chat_id=chat_id,
                    text=f"키워드 알림이 설정되었습니다.\n\n현재 저장된 알림 키워드:\n{updated_keywords_text}"
                )

        elif next_command == 'generate_chart':
            # 차트 생성 처리
            stock_list = [stock.strip() for stock in re.split('[,\n]', user_input) if stock.strip()]
            context.user_data['stock_list'] = stock_list
            context.user_data['generated_charts'] = []
            await process_generate_chart_stock_list(update, context, user_id, update.message)

        elif next_command == 'search_report':
            # 보고서 검색 처리
            stock_list = [stock.strip() for stock in re.split('[,\n]', user_input) if stock.strip()]
            context.user_data['stock_list'] = stock_list
            context.user_data['writeFromDate'] = (datetime.today() - timedelta(days=14)).strftime('%Y-%m-%d')
            context.user_data['writeToDate'] = datetime.today().strftime('%Y-%m-%d')
            await process_report_request(update, context, user_id, update.message)
            
        elif next_command == 'stock_quant':
            stock_list = [stock.strip() for stock in re.split('[,\n]', user_input) if stock.strip()]
            context.user_data['stock_list'] = stock_list

            all_quant_data = []
            print(stock_list)
            # 종목 검색
            for stock_name in stock_list:
                results = search_stock(stock_name)
                if results and len(results) == 1:
                    stock_name, stock_code = results[0]['name'], results[0]['code']
                    await update.message.reply_text(f"{stock_name} 퀀트 파일 생성 중입니다.")
                    quant_data = fetch_stock_info_quant_API(stock_code)
                    print(quant_data)
                    if quant_data:
                        all_quant_data.append(quant_data)
                elif results and len(results) > 1:
                    # 사용자에게 종목 선택을 받지 않고 종목명이 일치하는 값으로 자동 치환처리
                    select_stock = None  # 초기 값으로 None 설정
                    for result in results:
                        if result['name'] == stock_name:
                            select_stock = {'name': result['name'], 'code': result['code']}
                            break  # 조건을 만족하는 첫 번째 항목을 찾으면 루프 중단

                    if select_stock:
                        print('===>', select_stock)
                        stock_name, stock_code = select_stock['name'], select_stock['code']
                    else:
                        await update.message.reply_text(f"{stock_name} 을 찾을 수 없습니다.")
                        await update.message.reply_text(f"{results} \n 종목에서 ")
                        await update.message.reply_text(f"{results[0]['name']}로 처리됩니다. ")
                        stock_name, stock_code = results[0]['name'], results[0]['code']
                    
                    await update.message.reply_text(f"{stock_name} 퀀트 파일 생성 중입니다.")
                    quant_data = fetch_stock_info_quant_API(stock_code)

                    print(quant_data)
                    if quant_data:
                        all_quant_data.append(quant_data)
                    context.user_data['search_results'] = results
                else:
                    await update.message.reply_text(f"{stock_name} 검색 결과가 없습니다. 다시 시도하세요.")
            
            # Ensure the folder exists
            if not os.path.exists(EXCEL_FOLDER_PATH):
                os.makedirs(EXCEL_FOLDER_PATH)

            if all_quant_data:
                # Define the base file name and extension
                today_date = datetime.today().strftime('%y%m%d')
                base_file_name = f'stock_quant_{today_date}_{user_id}'
                file_extension = '.xlsx'
                counter = 0
                excel_file_name = os.path.join(EXCEL_FOLDER_PATH, f"{base_file_name}_{counter}{file_extension}")

                # Check if the file already exists and increment the sequence number if necessary
                while os.path.exists(excel_file_name):
                    counter += 1
                    excel_file_name = os.path.join(EXCEL_FOLDER_PATH, f"{base_file_name}_{counter}{file_extension}")

                # Create a DataFrame and save to Excel
                df = pd.DataFrame(all_quant_data)
                df.to_excel(excel_file_name, index=False, engine='openpyxl')

                # Send the file to the user
                if os.path.exists(excel_file_name):
                    with open(excel_file_name, 'rb') as file:
                        await context.bot.send_document(chat_id=chat_id, document=InputFile(file, filename=os.path.basename(excel_file_name)))
                else:
                    await context.bot.send_message(chat_id=chat_id, text="Excel 파일을 생성하는 데 문제가 발생했습니다.")
        
        else:
            # 업종 검색 처리
            upjong_list = fetch_upjong_list_API()
            upjong_map = {업종명: (등락률, 링크) for 업종명, 등락률, 링크 in upjong_list}
            upjong_number_map = {str(index + 1): 업종명 for index, (업종명, _, _) in enumerate(upjong_list)}

            if user_input in upjong_number_map:
                업종명 = upjong_number_map[user_input]
            else:
                업종명 = user_input

            if 업종명 in upjong_map:
                등락률, 링크 = upjong_map[업종명]
                await context.bot.send_message(chat_id=chat_id, text=f"입력한 업종명: {업종명}\n등락률: {등락률}")

                stock_info = fetch_stock_info_in_upjong(링크)
                if stock_info:
                    all_quant_data = []
                    for 종목명, _, _, _, 종목링크 in stock_info:
                        stock_code = 종목링크.split('=')[-1]
                        quant_data = fetch_stock_info_quant_API(stock_code)
                        if quant_data:
                            all_quant_data.append(quant_data)
                        else: pass

                    # Ensure the folder exists
                    if not os.path.exists(EXCEL_FOLDER_PATH):
                        os.makedirs(EXCEL_FOLDER_PATH)
                    today_date = datetime.today().strftime('%y%m%d')
                    excel_file_name = os.path.join(EXCEL_FOLDER_PATH, f'{업종명}_naver_quant_{today_date}.xlsx')

                    # Create a DataFrame and save to Excel with the sheet name as 업종명
                    df = pd.DataFrame(all_quant_data)
                    with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=업종명, index=False)

                    # 필터 기능 추가 및 하이퍼링크 처리
                    wb = load_workbook(excel_file_name)
                    ws = wb[업종명]
                        
                    # 필터 범위 지정
                    start_row, start_col = 1, 1
                    # end_row, end_col = df.shape[0] + 1, df.shape[1]
                    end_row = ws.max_row  # 실제 데이터가 있는 마지막 행을 가져옴
                    end_col = ws.max_column  # 실제 데이터가 있는 마지막 열을 가져옴
                    cell_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
                    ws.auto_filter.ref = cell_range  # 필터 범위 지정

                    # '네이버url' 열의 하이퍼링크 처리
                    if '네이버url' in df.columns:
                        nav_url_col_index = df.columns.get_loc('네이버url') + 1  # 1-based index
                        for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=nav_url_col_index, max_col=nav_url_col_index):
                            for cell in row:
                                if cell.value:
                                    cell.hyperlink = cell.value
                                    cell.font = Font(color="0000FF", underline="single")

                    # 엑셀 파일 저장
                    wb.save(excel_file_name)
                    wb.close()

                    print(f'퀀트 정보가 {excel_file_name} 파일에 저장되었습니다.')


                    if os.path.exists(excel_file_name):
                        with open(excel_file_name, 'rb') as file:
                            await context.bot.send_document(chat_id=chat_id, document=InputFile(file, filename=os.path.basename(excel_file_name)))
                    else:
                        await context.bot.send_message(chat_id=chat_id, text="Excel 파일을 생성하는 데 문제가 발생했습니다.")
                else:
                    await context.bot.send_message(chat_id=chat_id, text="종목 정보를 가져오는 데 문제가 발생했습니다.")
            else:
                await context.bot.send_message(chat_id=chat_id, text="입력한 업종명이 올바르지 않습니다.")
    
    except FileNotFoundError:
        await context.bot.send_message(chat_id=chat_id, text="파일이 존재하지 않습니다.")
    except IOError as e:
        await context.bot.send_message(chat_id=chat_id, text=f"파일 입출력 오류가 발생했습니다: {e}")
    except Exception as e:
        await context.bot.send_message(chat_id=chat_id, text=f"처리 중 오류가 발생했습니다: {e}")

# 파일 수신 및 시트별 데이터 출력
async def handle_document(update: Update, context: CallbackContext) -> None:
    chat_id = update.effective_chat.id
    document = update.message.document
    caption = update.message.caption
    next_command = context.user_data.get('next_command')

    if next_command == 'excel_quant' and document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        # 폴더가 존재하지 않으면 생성
        if not os.path.exists(EXCEL_FOLDER_PATH):
            os.makedirs(EXCEL_FOLDER_PATH)

        # 파일 다운로드
        file = await document.get_file()
        file_path = os.path.join(EXCEL_FOLDER_PATH, f"{chat_id}_{document.file_id}.xlsx")
        await file.download_to_drive(file_path)

        try:
            excel_data = pd.ExcelFile(file_path)
            sheet_names = excel_data.sheet_names
            update_message = ""

            # 파일 이름 및 경로 설정
            today_date = datetime.today().strftime('%y%m%d')
            counter = 0
            if not caption:
                caption = 'excel_quant'
            updated_file_name = os.path.join(EXCEL_FOLDER_PATH, f'{caption}_{chat_id}_{today_date}_{counter}.xlsx')
            while os.path.exists(updated_file_name):
                counter += 1
                updated_file_name = os.path.join(EXCEL_FOLDER_PATH, f'{caption}_{chat_id}_{today_date}_{counter}.xlsx')

            # 최초 메시지 전송
            message = await context.bot.send_message(chat_id=chat_id, text="처리 중...")
            with pd.ExcelWriter(updated_file_name, engine='openpyxl') as writer:
                for sheet_name in sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    update_message += f"============\n시트 이름: {sheet_name}\n"

                    stock_update_count = 0  # 갱신된 종목 수를 세기 위한 변수
                    temp_update_messages = []  # 갱신 중 메시지를 저장할 리스트

                    for index, row in df.iterrows():
                        naver_url = row.get('네이버url')
                        stock_code = row.get('종목코드')
                        stock_name = row.get('종목명')
                        memo = row.get('비고(메모)')

                        # 빈 값 처리
                        naver_url = '' if pd.isna(naver_url) else naver_url
                        stock_code = '' if pd.isna(stock_code) else stock_code
                        stock_name = '' if pd.isna(stock_name) else stock_name
                        memo = '' if pd.isna(memo) else memo

                        if naver_url:
                            stock_code = naver_url.replace('https://finance.naver.com/item/main.naver?code=', '')

                        # 퀀트 데이터 가져오기
                        if stock_code:
                            stock_info = search_stock(stock_code)
                            quant_data = fetch_stock_info_quant_API(stock_info[0]['code'])
                        elif stock_name:
                            stock_info = search_stock(stock_name)
                            quant_data = fetch_stock_info_quant_API(stock_info[0]['code'])
                        else:
                            quant_data = None

                        # 각 종목 갱신 메시지 업데이트
                        if quant_data:
                            for key, value in quant_data.items():
                                if key == '비고(메모)': value = row.get('비고(메모)')
                                elif key == '분류': value = row.get('분류')
                                df.at[index, key] = value

                            stock_update_count += 1  # 갱신된 종목 수 증가
                            
                            # # 메시지 수정
                            # temp_update_messages.append(f"   [{stock_name}] 퀀트 데이터 갱신 중..\n")
                            
                            # try:
                            #     # 메시지 간에 1초 대기
                            #     await asyncio.sleep(1.5)
                            #     await context.bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=update_message + ''.join(temp_update_messages))
                            # except Exception as e:
                            #     print(f"메시지 수정 중 오류 발생: {e}")

                    # 시트 갱신 완료 메시지 추가
                    if stock_update_count > 0:
                        update_message += f"   {stock_update_count}종목 갱신 완료\n"
                    else:
                        update_message += "   갱신된 종목이 없습니다.\n"

                    # 시트가 넘어갈 때 메시지 업데이트
                    try:
                        await asyncio.sleep(1.5)
                        await context.bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=update_message)
                    except Exception as e:
                        print(f"메시지 수정 중 오류 발생: {[{stock_name}]}종목을 확인 하세요 \n{e}")

                    # 갱신된 시트를 새로운 엑셀 파일에 저장
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 필터 및 하이퍼링크 처리
            wb = load_workbook(updated_file_name)
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                
                start_row, start_col = 1, 1
                end_row = ws.max_row  # 실제 데이터가 있는 마지막 행을 가져옴
                end_col = ws.max_column  # 실제 데이터가 있는 마지막 열을 가져옴
                cell_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
                ws.auto_filter.ref = cell_range

                for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=1, max_col=end_col):
                    for cell in row:
                        if cell.value and '네이버url' in df.columns and cell.column == df.columns.get_loc('네이버url') + 1:
                            cell.hyperlink = cell.value
                            cell.font = Font(color="0000FF", underline="single")

            wb.save(updated_file_name)
            wb.close()

            # 모든 작업이 완료된 후 최종 메시지 수정
            update_message += "============\n엑셀 데이터 전송 완료"
            # 메시지 간에 1초 대기
            await asyncio.sleep(1.5)
            if len(update_message) > 3000:
                # 메시지가 3000자를 넘는 경우, 메시지 분할 전송
                for i in range(0, len(update_message), 3000):
                    await context.bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=update_message[i:i + 3000])
                    await asyncio.sleep(1)
            else:
                await context.bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=update_message)

            # 파일 전송
            with open(updated_file_name, 'rb') as file:
                # 메시지 간에 1초 대기
                await asyncio.sleep(1.5)
                await context.bot.send_document(chat_id=chat_id, document=InputFile(file, filename=updated_file_name))

        except Exception as e:
            await context.bot.send_message(chat_id=chat_id, text=f"{stock_name}\n처리 중 오류가 발생했습니다: {e}")

        context.user_data['next_command'] = None
    else:
        await context.bot.send_message(chat_id=chat_id, text="올바른 엑셀 파일을 전송해 주세요.")

# 엑셀 셀 주소 변환 함수
def number_to_coordinate(rc):
    row_idx, col_idx = rc[0], rc[1]
    col_string = get_column_letter(col_idx)
    return f'{col_string}{row_idx}'

def main():
    load_dotenv()  # .env 파일의 환경 변수를 로드합니다
    env = os.getenv('ENV')
    print(env)
    if env == 'production':
        token = os.getenv('TELEGRAM_BOT_TOKEN_PROD')
    else:
        token = os.getenv('TELEGRAM_BOT_TOKEN_TEST')

    application = ApplicationBuilder().token(token).build()

    recent_searches = load_recent_searches()
    application.bot_data['recent_searches'] = recent_searches

    application.add_handler(CommandHandler("generate_chart", generate_chart))  # /generate_chart 명령어 추가
    application.add_handler(CommandHandler("recent", show_recent_searches))  # 최근 검색 종목 명령어 추가
    application.add_handler(CommandHandler("search_report", search_report))  # 레포트 검색기 명령어 추가
    application.add_handler(CommandHandler("naver_upjong_quant", show_upjong_list))  # 업종 목록 표시
    application.add_handler(CommandHandler("stock_quant", stock_quant))  
    application.add_handler(CommandHandler("excel_quant", excel_quant))
    application.add_handler(CommandHandler("report_alert_keyword", report_alert_keyword))  # 알림 키워드 명령어 추가


    application.add_handler(CallbackQueryHandler(route_command_based_on_user_input, pattern=r'^\d{6}$'))
    application.add_handler(CallbackQueryHandler(previous_search, pattern='^previous_search$'))

    # Add message handlers
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # asyncio 이벤트 루프에서 명령어 설정
    loop = asyncio.get_event_loop()
    loop.run_until_complete(set_commands(application.bot))

    application.run_polling()


if __name__ == '__main__':
    if not os.path.exists(CHART_DIR):
        os.makedirs(CHART_DIR)
    main()